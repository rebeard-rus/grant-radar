#!/usr/bin/env python3
"""
Грант-радар: обходит источники (страницы фондов и Telegram-каналы), находит
НОВЫЕ гранты/конкурсы по ключевым словам и присылает их в Telegram.

Память о том, что уже показывал, лежит в seen.json (workflow коммитит его
обратно в репозиторий).

Почему так устроено:
  • Многие сайты фондов рендерятся через JS или отдают 403 не-браузеру
    (rscf.ru/contests, fasie.ru), поэтому надёжнее брать:
      – статичные страницы-новости (type: html) — диффим по тексту заголовков;
      – Telegram-каналы через веб-превью t.me/s/<канал> (type: telegram) —
        чистый HTML, у каждого поста есть уникальная ссылка (по ней и диффим).
  • Не завязываемся на хрупкие CSS-селекторы конкретных сайтов.
"""

import datetime
import hashlib
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).parent
CONFIG_PATH = ROOT / "config.json"
SEEN_PATH = ROOT / "seen.json"
PENDING_PATH = ROOT / "pending.json"  # очередь находок для режима digest
ARCHIVE_PATH = ROOT / "archive.json"  # исторический архив всех находок
REPORT_PATH = ROOT / "archive.md"     # человекочитаемый отчёт (--report)
BOT_STATE_PATH = ROOT / "bot_state.json"  # offset getUpdates для бота-поиска

# «Браузерный» набор заголовков — часть сайтов иначе отвечает 403.
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9",
}

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "").strip()


def load_json(path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            print(f"⚠️  Не смог прочитать {path.name}, использую default")
    return default


def norm(text):
    return re.sub(r"\s+", " ", text.lower()).strip()


def md5(s):
    return hashlib.md5(s.encode("utf-8")).hexdigest()


def fetch(url):
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or resp.encoding
    return resp.text


def clean(text):
    """Схлопываем спам-повторы эмодзи ("🎤 🎤 🎤" -> "🎤") и лишние пробелы."""
    text = re.sub(r"(\S+)(?:\s+\1){2,}", r"\1", text)   # повторы токенов
    text = re.sub(r"(\S)\1{2,}", r"\1", text)           # повторы одного символа
    return " ".join(text.split()).strip()


def shorten(text, limit):
    text = clean(text)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def collect_html(html, base_url, cfg):
    """Заголовки со статичной страницы (ссылки, h1..h5, li).

    Возвращает список item-словарей: title(показ), text(для фильтра),
    link, key(для дедупа — по нормализованному тексту).
    """
    soup = BeautifulSoup(html, "html.parser")
    min_len = cfg.get("min_title_len", 15)
    max_len = cfg.get("max_title_len", 240)

    seen_norm = set()
    items = []
    for tag in soup.find_all(["a", "h1", "h2", "h3", "h4", "h5", "li"]):
        title = " ".join(tag.get_text(" ", strip=True).split())
        if not (min_len <= len(title) <= max_len):
            continue
        n = norm(title)
        if n in seen_norm:
            continue
        seen_norm.add(n)
        href = tag.get("href") if tag.name == "a" else None
        if not href:
            a = tag.find("a", href=True)
            href = a["href"] if a else None
        link = urljoin(base_url, href) if href else base_url
        items.append({"title": title, "text": title, "link": link, "key": n})
    return items


def collect_telegram(html, cfg):
    """Посты Telegram-канала из веб-превью t.me/s/<канал>.

    Для фильтра берём ПОЛНЫЙ текст поста, для показа — укороченный.
    Дедуп — по уникальной ссылке на пост.
    """
    soup = BeautifulSoup(html, "html.parser")
    limit = cfg.get("max_title_len", 240)
    items = []
    for msg in soup.select("div.tgme_widget_message"):
        body = msg.select_one("div.tgme_widget_message_text")
        if not body:
            continue
        text = body.get_text(" ", strip=True)
        if len(text) < cfg.get("min_title_len", 15):
            continue
        post = msg.get("data-post")  # вида "channel/123"
        link = f"https://t.me/{post}" if post else "https://t.me/"
        items.append(
            {"title": shorten(text, limit), "text": text, "link": link, "key": link}
        )
    return items


def matches(text, cfg):
    low = text.lower()
    if not any(k.lower() in low for k in cfg["keywords"]):
        return False
    if cfg.get("require_signal_word", True):
        return any(s.lower() in low for s in cfg.get("signal_words", []))
    return True


def matched_keywords(text, cfg):
    """Какие ключевые слова (направления) сработали в тексте."""
    low = text.lower()
    return [k for k in cfg["keywords"] if k.lower() in low]


def today_iso():
    return datetime.date.today().isoformat()


def _tg_post(method, data):
    """POST в Telegram API с ретраем на 429 (rate limit)."""
    url = f"https://api.telegram.org/bot{TOKEN}/{method}"
    r = None
    for _ in range(3):
        r = requests.post(url, data=data, timeout=35)
        if r.status_code == 429:
            wait = r.json().get("parameters", {}).get("retry_after", 2)
            time.sleep(min(wait, 30))
            continue
        return r
    return r


def _strip_html(text):
    return re.sub(r"<[^>]+>", "", text)


def tg_send(text, chat_id=None, reply_markup=None):
    target = chat_id or CHAT_ID
    if not (TOKEN and target):
        print("⚠️  Нет TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID — в консоль:\n"
              + _strip_html(text) + "\n")
        return
    data = {
        "chat_id": target,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": "true",
    }
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup, ensure_ascii=False)
    r = _tg_post("sendMessage", data)
    if r is not None and not r.ok:
        # Битый HTML? Повторяем как обычный текст, чтобы не потерять сообщение.
        data["text"] = _strip_html(text)
        data.pop("parse_mode", None)
        r = _tg_post("sendMessage", data)
        if r is not None and not r.ok:
            print(f"⚠️  Telegram вернул {r.status_code}: {r.text[:200]}")


def tg_edit_message(chat_id, message_id, text, reply_markup=None):
    """Редактирует существующее сообщение (для пагинации «показать ещё»)."""
    data = {
        "chat_id": chat_id,
        "message_id": message_id,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": "true",
    }
    if reply_markup is not None:
        data["reply_markup"] = json.dumps(reply_markup, ensure_ascii=False)
    r = _tg_post("editMessageText", data)
    # "message is not modified" — не ошибка; на прочие тихо забиваем.
    return r is not None and r.ok


def tg_answer_callback(callback_id):
    """Гасит «часики» на нажатой кнопке."""
    try:
        _tg_post("answerCallbackQuery", {"callback_query_id": callback_id})
    except Exception:
        pass


def esc(s):
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def scan(cfg):
    """Обходит все источники и возвращает список (src, item) по фильтру."""
    found = []
    for src in cfg["sources"]:
        try:
            html = fetch(src["url"])
        except Exception as e:
            print(f"⚠️  {src['name']}: не открылся ({e})")
            continue
        if src.get("type") == "telegram":
            items = collect_telegram(html, cfg)
        else:
            items = collect_html(html, src["url"], cfg)
        hits = [it for it in items if matches(it["text"], cfg)]
        print(f"• {src['name']}: элементов {len(items)}, по фильтру {len(hits)}")
        for it in hits:
            found.append((src, it))
    return found


def stream_item_msg(src, it):
    return (
        f"{src.get('emoji', '📌')} <b>{esc(it['title'])}</b>\n"
        f"<i>{esc(src['name'])}</i>\n{esc(it['link'])}"
    )


TG_LIMIT = 3500  # с запасом от лимита Telegram в 4096 символов


def build_digest(pending):
    """Собирает накопленные находки в дайджест-сообщения (с разбивкой по лимиту).

    Группирует по источнику; возвращает список готовых HTML-строк.
    """
    by_source = {}
    order = []
    for it in pending:
        name = it["name"]
        if name not in by_source:
            by_source[name] = {"emoji": it.get("emoji", "📌"), "items": []}
            order.append(name)
        by_source[name]["items"].append(it)

    header = f"🗂 <b>Грант-дайджест</b> — {len(pending)} новых за период\n"
    blocks = []
    for name in order:
        grp = by_source[name]
        lines = [f"\n{grp['emoji']} <b>{esc(name)}</b>"]
        for it in grp["items"]:
            lines.append(f"• <a href=\"{esc(it['link'])}\">{esc(it['title'])}</a>")
        blocks.append("\n".join(lines))

    # Склеиваем блоки в сообщения, не превышая лимит.
    messages = []
    cur = header
    for block in blocks:
        if len(cur) + len(block) > TG_LIMIT and cur.strip() != header.strip():
            messages.append(cur.rstrip())
            cur = header + block
        else:
            cur += block
    if cur.strip():
        messages.append(cur.rstrip())
    return messages


def flush_digest(cfg):
    """Шлёт накопленный дайджест и очищает очередь (команда --flush)."""
    pending = load_json(PENDING_PATH, [])
    if not pending:
        tg_send("🗂 За прошедший период новых подходящих конкурсов не появилось.")
        print("Очередь пуста — дайджест не отправлен.")
        return
    for msg in build_digest(pending):
        tg_send(msg)
        time.sleep(0.4)
    PENDING_PATH.write_text("[]\n", encoding="utf-8")
    print(f"✅ Дайджест отправлен: {len(pending)} позиций, очередь очищена.")


def chunk_messages(lines, limit=TG_LIMIT):
    """Склеивает строки в сообщения, не превышающие лимит Telegram."""
    messages = []
    cur = ""
    for ln in lines:
        piece = ("\n" if cur else "") + ln
        if cur and len(cur) + len(piece) > limit:
            messages.append(cur)
            cur = ln
        else:
            cur += piece
    if cur:
        messages.append(cur)
    return messages


# ─────────────────────────  Архив  ─────────────────────────

def archive_upsert(archive, src, it, cfg, today):
    """Добавляет/обновляет запись в историческом архиве (по ключу)."""
    kws = matched_keywords(it["text"], cfg)
    rec = archive.get(it["key"])
    if rec:
        rec["title"] = it["title"]
        rec["link"] = it["link"]
        rec["source"] = src["name"]
        rec["emoji"] = src.get("emoji", "📌")
        rec["keywords"] = sorted(set(rec.get("keywords", [])) | set(kws))
        rec["last_seen"] = today
    else:
        archive[it["key"]] = {
            "title": it["title"],
            "link": it["link"],
            "source": src["name"],
            "emoji": src.get("emoji", "📌"),
            "keywords": kws,
            "first_seen": today,
            "last_seen": today,
        }


def search_archive(archive, query, cfg):
    """Ищет по архиву: заголовок, источник и направления (ключевые слова).

    Гибкое совпадение — запрос как подстрока заголовка/источника, а по
    направлениям срабатывает и «горнодобыча» → корень «горн», и наоборот.
    """
    q = norm(query)
    if not q:
        return []
    tokens = q.split()
    results = []
    for rec in archive.values():
        hay = norm(rec.get("title", "") + " " + rec.get("source", ""))
        kws = [k.lower() for k in rec.get("keywords", [])]
        hit = q in hay or all(
            (t in hay) or any(t in kw or kw in t for kw in kws) for t in tokens
        )
        if hit:
            results.append(rec)
    results.sort(key=lambda r: r.get("first_seen", ""), reverse=True)
    return results


# ─────────────  Импорт внешних баз грантов (--import)  ─────────────
#
# Друг может «подгрузить» свои исторические базы: CSV / Excel (.xlsx) / JSON.
# Колонки распознаются по названиям (рус/англ, регистр не важен) — см. ниже.
# Записи вливаются в тот же archive.json, и бот сразу ищет по ним.

COLUMN_ALIASES = {
    "title": ["title", "название", "наименование", "заголовок", "name",
              "грант", "конкурс", "программа", "тема гранта"],
    "link": ["link", "url", "ссылка", "сайт", "href", "страница"],
    "source": ["source", "источник", "организация", "фонд", "оператор",
               "grantor", "организатор", "донор"],
    "date": ["date", "дата", "год", "year", "first_seen", "дата начала",
             "опубликовано", "дата публикации"],
    "description": ["description", "описание", "аннотация", "summary",
                    "details", "текст", "суть"],
    "amount": ["amount", "сумма", "размер", "финансирование", "бюджет",
               "размер гранта", "сумма гранта"],
    "deadline": ["deadline", "срок", "дедлайн", "прием до", "приём до",
                 "окончание", "срок подачи"],
    "category": ["category", "категория", "направление", "тема", "область",
                 "тематика", "keywords", "теги", "сфера"],
}


def _norm_key(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _row_get(row, canonical):
    for alias in COLUMN_ALIASES.get(canonical, []):
        val = row.get(alias)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _read_rows(path):
    """Читает CSV / JSON / XLSX в список словарей с нормализованными ключами."""
    p = Path(path)
    ext = p.suffix.lower()
    rows = []
    if ext == ".json":
        data = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = list(data.values())
        for item in data:
            if isinstance(item, dict):
                rows.append({_norm_key(k): v for k, v in item.items()})
    elif ext == ".csv":
        import csv
        text = p.read_text(encoding="utf-8-sig")
        head = text[:2000]
        delim = ";" if head.count(";") > head.count(",") else ","
        for r in csv.DictReader(text.splitlines(), delimiter=delim):
            rows.append({_norm_key(k): v for k, v in r.items() if k})
    elif ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Для .xlsx нужен openpyxl: pip install openpyxl "
                     "(или сохрани файл как CSV).")
        ws = load_workbook(p, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        headers = [_norm_key(h) for h in next(it, [])]
        for vals in it:
            rows.append({h: v for h, v in zip(headers, vals) if h})
    else:
        sys.exit(f"Формат {ext} не поддержан. Используй .csv, .json или .xlsx.")
    return rows


def _parse_date(raw):
    """Пытается привести дату к ISO (YYYY-MM-DD); иначе — как есть."""
    raw = str(raw or "").strip()
    if not raw:
        return ""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if m:
        return m.group(0)
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", raw)
    if m:
        d, mth, y = m.groups()
        return f"{y}-{int(mth):02d}-{int(d):02d}"
    m = re.search(r"\b(?:19|20)\d{2}\b", raw)
    if m:
        return f"{m.group(0)}-01-01"
    return raw


def import_file(path, archive, cfg):
    """Вливает один файл в архив. Возвращает (добавлено, обновлено, пропущено)."""
    rows = _read_rows(path)
    added = updated = skipped = 0
    default_source = f"Импорт: {Path(path).stem}"
    for row in rows:
        title = _row_get(row, "title")
        if not title:
            skipped += 1
            continue
        link = _row_get(row, "link")
        source = _row_get(row, "source") or default_source
        date = _parse_date(_row_get(row, "date")) or today_iso()
        desc = _row_get(row, "description")
        category = _row_get(row, "category")
        kws = matched_keywords(f"{title} {desc} {category}", cfg)
        key = "imp:" + md5(link or _norm_key(title))
        exists = key in archive
        rec = archive.get(key, {})
        rec.update({
            "title": title[:400],
            "link": link,
            "source": source,
            "emoji": rec.get("emoji", "📁"),
            "keywords": sorted(set(rec.get("keywords", [])) | set(kws)),
            "first_seen": rec.get("first_seen", date),
            "last_seen": date,
            "origin": "import",
        })
        if desc:
            rec["description"] = desc[:800]
        for opt in ("amount", "deadline"):
            v = _row_get(row, opt)
            if v:
                rec[opt] = v[:200]
        archive[key] = rec
        updated += exists
        added += not exists
    return added, updated, skipped


def cmd_import(cfg, paths):
    """Импорт файлов/папок с историческими грантами в archive.json."""
    exts = (".csv", ".json", ".xlsx", ".xlsm")
    targets = []
    for pth in paths:
        p = Path(pth)
        if p.is_dir():
            targets += [str(f) for f in sorted(p.iterdir()) if f.suffix.lower() in exts]
        elif p.exists():
            targets.append(str(p))
        else:
            print(f"⚠️  не найдено: {pth}")
    if not targets:
        sys.exit("Нет файлов для импорта (.csv/.json/.xlsx или папка с ними).")
    archive = load_json(ARCHIVE_PATH, {})
    before = len(archive)
    for t in targets:
        try:
            a, u, s = import_file(t, archive, cfg)
            print(f"📥 {Path(t).name}: +{a} новых, {u} обновлено, {s} пропущено")
        except Exception as e:
            print(f"⚠️  {Path(t).name}: не смог импортировать ({e})")
    ARCHIVE_PATH.write_text(
        json.dumps(archive, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    print(f"💾 archive.json: было {before} → стало {len(archive)} записей.")


# ─────────────────────  Отчёт (--report)  ─────────────────────

def build_report(archive, cfg):
    """Markdown-отчёт, сгруппированный по направлениям (для Ctrl+F)."""
    by_kw = {}
    for rec in archive.values():
        for kw in rec.get("keywords", []) or ["(без направления)"]:
            by_kw.setdefault(kw, []).append(rec)

    order = sorted(by_kw, key=lambda k: (-len(by_kw[k]), k))
    out = [
        "# 🗂 Архив грантов (грант-радар)",
        "",
        f"Всего записей: **{len(archive)}**. Обновлено: {today_iso()}.",
        "",
        "## Направления",
        "",
    ]
    for kw in order:
        anchor = re.sub(r"[^\w]+", "-", kw.strip()).strip("-").lower()
        out.append(f"- [{kw}](#{anchor}) — {len(by_kw[kw])}")
    out.append("")

    for kw in order:
        out.append(f"## {kw}")
        out.append("")
        recs = sorted(by_kw[kw], key=lambda r: r.get("first_seen", ""), reverse=True)
        for r in recs:
            date = r.get("first_seen", "?")
            out.append(f"- {date} — [{r['title']}]({r['link']}) · {r.get('source','')}")
        out.append("")
    return "\n".join(out)


def cmd_report(cfg):
    archive = load_json(ARCHIVE_PATH, {})
    if not archive:
        print("⚠️  Архив пуст — сначала запусти обход (python radar.py).")
        return
    REPORT_PATH.write_text(build_report(archive, cfg), encoding="utf-8")
    print(f"📄 Отчёт записан: {REPORT_PATH.name} ({len(archive)} записей)")


# ────────────────  Поиск: локально и через бот  ────────────────

PAGE_DEFAULT = 5
QUERY_CACHE = {}  # md5(query)[:8] -> query, для пагинации текстовых запросов

MONTHS_RU = ["", "января", "февраля", "марта", "апреля", "мая", "июня",
             "июля", "августа", "сентября", "октября", "ноября", "декабря"]
NUM_EMOJI = ["0️⃣", "1️⃣", "2️⃣", "3️⃣", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]


def human_date(iso, with_year=False):
    """'2026-07-05' -> '5 июля' (или '5 июля 2026')."""
    try:
        d = datetime.date.fromisoformat(str(iso)[:10])
    except (ValueError, TypeError):
        return str(iso or "")
    s = f"{d.day} {MONTHS_RU[d.month]}"
    return f"{s} {d.year}" if with_year else s


def is_fresh(iso, days=7):
    try:
        d = datetime.date.fromisoformat(str(iso)[:10])
    except (ValueError, TypeError):
        return False
    return 0 <= (datetime.date.today() - d).days <= days


def num_badge(n):
    return NUM_EMOJI[n] if 0 <= n <= 10 else f"{n}."


def archive_start(archive):
    dates = [r.get("first_seen") for r in archive.values() if r.get("first_seen")]
    return human_date(min(dates), with_year=True) if dates else "сегодня"


def direction_counts(cfg, archive):
    """[(index, direction, count)] — сколько грантов по каждому направлению."""
    out = []
    for i, d in enumerate(cfg.get("directions", [])):
        n = len(search_archive(archive, d.get("query", d.get("label", "")), cfg))
        out.append((i, d, n))
    return out


def format_card(n, rec):
    title = (rec.get("title") or "").strip() or "(без названия)"
    date = human_date(rec.get("first_seen"))
    fresh = " 🆕" if is_fresh(rec.get("first_seen")) else ""
    meta = " · ".join(x for x in [f"{rec.get('emoji', '📌')} {esc(rec.get('source', ''))}",
                                  date] if x.strip())
    lines = [f"{num_badge(n)} <b>{esc(title)}</b>", f"   {meta}{fresh}"]
    if rec.get("link"):
        lines.append(f"   🔗 <a href=\"{esc(rec['link'])}\">Открыть</a>")
    return "\n".join(lines)


def page_keyboard(page_cb, offset, page, total, extra_top=None):
    rows = []
    if extra_top:
        rows.append(extra_top)
    nav = []
    if offset > 0:
        nav.append({"text": "⬆️ Назад",
                    "callback_data": f"{page_cb}:{max(0, offset - page)}"})
    if offset + page < total:
        nav.append({"text": f"⬇️ Ещё {min(page, total - offset - page)}",
                    "callback_data": f"{page_cb}:{offset + page}"})
    if nav:
        rows.append(nav)
    rows.append([{"text": "⬅️ К направлениям", "callback_data": "home"}])
    return {"inline_keyboard": rows}


def render_page(title, emoji, results, offset, page_cb, cfg, extra_top=None):
    page = int(cfg.get("bot_page_size", PAGE_DEFAULT))
    total = len(results)
    offset = max(0, min(offset, (total - 1) // page * page if total else 0))
    chunk = results[offset:offset + page]
    head = (f"{emoji} <b>{esc(title)}</b>\n"
            f"Найдено: <b>{total}</b> · показываю {offset + 1}–{offset + len(chunk)}\n\n")
    body = "\n\n".join(format_card(offset + i + 1, r) for i, r in enumerate(chunk))
    return {"text": head + body,
            "markup": page_keyboard(page_cb, offset, page, total, extra_top)}


def render_empty(query, cfg, archive):
    text = (
        f"📭 По запросу «{esc(query)}» в архиве пока пусто.\n"
        f"Веду архив с {archive_start(archive)} — тема появится, "
        "как только встретится в источниках.\n\n"
        "👇 Загляни в соседние направления:"
    )
    return {"text": text, "markup": directions_keyboard(cfg, archive)}


def render_query(query, offset, cfg, archive, title=None, emoji="🔍", page_cb=None):
    results = search_archive(archive, query, cfg)
    if not results:
        return render_empty(query, cfg, archive)
    if page_cb is None:
        h = md5(query)[:8]
        QUERY_CACHE[h] = query
        page_cb = f"q:{h}"
    return render_page(title or query, emoji, results, offset, page_cb, cfg)


def render_fresh(days, offset, cfg, archive):
    results = sorted(
        (r for r in archive.values() if is_fresh(r.get("first_seen"), days)),
        key=lambda r: r.get("first_seen", ""), reverse=True,
    )
    periods = [
        {"text": ("• 7 дней" if days == 7 else "7 дней"), "callback_data": "f:7:0"},
        {"text": ("• 30 дней" if days == 30 else "30 дней"), "callback_data": "f:30:0"},
    ]
    if not results:
        text = (f"🆕 За {days} дн. новых грантов в архиве нет.\n"
                f"Архив веду с {archive_start(archive)}.")
        return {"text": text, "markup": {"inline_keyboard": [
            periods, [{"text": "⬅️ К направлениям", "callback_data": "home"}]]}}
    return render_page(f"Свежие за {days} дн.", "🆕", results, offset,
                       f"f:{days}", cfg, extra_top=periods)


def render_stats(cfg, archive):
    counts = sorted(direction_counts(cfg, archive), key=lambda t: -t[2])
    lines = [f"📊 <b>Архив: {len(archive)} грантов</b> · с {archive_start(archive)}", ""]
    for _, d, n in counts:
        if n:
            lines.append(f"{d['label']} — <b>{n}</b>")
    empty = [d["label"] for _, d, n in counts if not n]
    if empty:
        lines += ["", "<i>Пока пусто: " + ", ".join(empty) + "</i>"]
    return {"text": "\n".join(lines), "markup": {"inline_keyboard": [
        [{"text": "🆕 Свежие", "callback_data": "f:7:0"}],
        [{"text": "⬅️ К направлениям", "callback_data": "home"}]]}}


def render_start(cfg, archive):
    n = len(archive)
    ndir = sum(1 for _, _, c in direction_counts(cfg, archive) if c)
    text = (
        "👋 <b>Привет! Я — Грант-радар.</b>\n\n"
        f"Слежу за источниками грантов ({len(cfg.get('sources', []))} шт.: РНФ, "
        "Фонд Бортника, Grantfull и др.) и веду по ним архив конкурсов.\n\n"
        f"📚 Сейчас в архиве <b>{n}</b> грантов"
        + (f" по <b>{ndir}</b> направлениям" if ndir else "")
        + f" · с {archive_start(archive)}.\n\n"
        "Что я умею:\n"
        "🔍 Напиши тему — найду гранты: <code>биотех</code>, <code>станки</code>\n"
        "🧭 Или жми кнопку направления ниже\n"
        "🆕 /fresh — что появилось недавно\n"
        "📊 /stats — что есть в архиве\n"
        "❓ /help — подсказка\n\n"
        "👇 С чего начнём?"
    )
    return {"text": text, "markup": directions_keyboard(cfg, archive)}


def render_help(cfg, archive):
    text = (
        "🛰 <b>Грант-радар — поиск по архиву</b>\n\n"
        "• Напиши тему словом: <code>горн</code>, <code>малое предприятие</code>, "
        "<code>искусственный интеллект</code>.\n"
        "• Или нажми кнопку направления.\n"
        "• Листай результаты кнопкой «⬇️ Ещё».\n\n"
        "Команды: /start · /fresh · /stats"
    )
    return {"text": text, "markup": directions_keyboard(cfg, archive)}


def directions_keyboard(cfg, archive, expanded=False):
    """Клавиатура направлений: счётчики, сортировка по наполненности, топ + «Ещё»."""
    if not cfg.get("directions"):
        counts = {}
        for r in archive.values():
            for k in r.get("keywords", []):
                counts[k] = counts.get(k, 0) + 1
        top = sorted(counts, key=lambda k: (-counts[k], k))[:12]
        cfg = dict(cfg, directions=[{"label": k, "query": k} for k in top])
    if not cfg.get("directions"):
        return None

    triples = sorted(direction_counts(cfg, archive), key=lambda t: (-t[2], t[0]))
    top_n = len(triples) if expanded else int(cfg.get("bot_dir_buttons", 6))
    rows, row = [], []
    for i, d, n in triples[:top_n]:
        label = d["label"] + (f" · {n}" if n else "")
        row.append({"text": label, "callback_data": f"d:{i}:0"})
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    if not expanded and len(triples) > top_n:
        rows.append([{"text": "🧭 Ещё направления ▸", "callback_data": "more"}])
    rows.append([
        {"text": "🆕 Свежие", "callback_data": "f:7:0"},
        {"text": "📊 Статистика", "callback_data": "stats"},
    ])
    return {"inline_keyboard": rows}


def handle_message(text, cfg, archive):
    """Текстовое сообщение/команда -> один render-словарь {text, markup}."""
    q = (text or "").strip()
    if q.startswith("/"):
        parts = q.split(maxsplit=1)
        cmd = parts[0].lower().split("@")[0]
        arg = parts[1].strip() if len(parts) > 1 else ""
        if cmd == "/start":
            return render_start(cfg, archive)
        if cmd in ("/help", "/?"):
            return render_help(cfg, archive)
        if cmd in ("/stats", "/статистика"):
            return render_stats(cfg, archive)
        if cmd in ("/fresh", "/свежие", "/new"):
            return render_fresh(7, 0, cfg, archive)
        if cmd in ("/directions", "/napravleniya"):
            return render_start(cfg, archive)
        q = arg  # /grants, /search, /find ... — остаток как запрос
    if not q:
        return render_start(cfg, archive)
    return render_query(q, 0, cfg, archive)


def handle_callback(data, cfg, archive):
    """Нажатие кнопки -> render-словарь для редактирования сообщения."""
    data = data or ""
    if data == "home":
        return render_start(cfg, archive)
    if data == "more":
        return {"text": "🧭 <b>Все направления</b>\nВыбери тему:",
                "markup": directions_keyboard(cfg, archive, expanded=True)}
    if data == "stats":
        return render_stats(cfg, archive)
    parts = data.split(":")
    kind = parts[0]
    if kind == "d" and len(parts) == 3 and parts[1].isdigit():
        idx = int(parts[1])
        off = int(parts[2]) if parts[2].isdigit() else 0
        dirs = cfg.get("directions", [])
        if 0 <= idx < len(dirs):
            d = dirs[idx]
            return render_query(d.get("query", d["label"]), off, cfg, archive,
                                title=d["label"], emoji="🎯", page_cb=f"d:{idx}")
    if kind == "f" and len(parts) >= 2 and parts[1].isdigit():
        days = int(parts[1])
        off = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        return render_fresh(days, off, cfg, archive)
    if kind == "q" and len(parts) == 3:
        query = QUERY_CACHE.get(parts[1])
        off = int(parts[2]) if parts[2].isdigit() else 0
        if query:
            return render_query(query, off, cfg, archive, page_cb=f"q:{parts[1]}")
        return {"text": "⏳ Запрос устарел. Напиши тему заново или выбери направление 👇",
                "markup": directions_keyboard(cfg, archive)}
    if data.startswith("d:"):  # старый формат кнопок "d:<текст>"
        return render_query(data[2:], 0, cfg, archive)
    return render_start(cfg, archive)


def cmd_search(cfg, query):
    """Локальный поиск: печатает результаты в консоль (без Telegram)."""
    archive = load_json(ARCHIVE_PATH, {})
    results = search_archive(archive, query, cfg)
    if not results:
        print(f"Ничего не найдено по «{query}». В архиве {len(archive)} записей.")
        return
    print(f"🔍 «{query}» — найдено {len(results)}:\n")
    for i, r in enumerate(results[:30], 1):
        print(f"{i}. {r.get('title', '')}")
        print(f"   {r.get('source', '')} · {human_date(r.get('first_seen'), True)}")
        if r.get("link"):
            print(f"   {r['link']}")
        print()


def tg_get_updates(offset, timeout):
    r = requests.get(
        f"https://api.telegram.org/bot{TOKEN}/getUpdates",
        params={"offset": offset, "timeout": timeout},
        timeout=timeout + 10,
    )
    r.raise_for_status()
    return r.json().get("result", [])


def process_update(up, cfg, archive):
    """Обрабатывает один апдейт: нажатие кнопки (редактирует) или сообщение."""
    cq = up.get("callback_query")
    if cq:
        render = handle_callback(cq.get("data", ""), cfg, archive)
        msg = cq.get("message") or {}
        chat = msg.get("chat", {}).get("id")
        mid = msg.get("message_id")
        if render and chat and mid:
            ok = tg_edit_message(chat, mid, render["text"], render.get("markup"))
            if not ok:  # сообщение слишком старое для правки — шлём новое
                tg_send(render["text"], chat_id=chat, reply_markup=render.get("markup"))
        tg_answer_callback(cq["id"])
        return
    msg = up.get("message")
    if not msg or not msg.get("text"):
        return  # игнорируем стикеры/фото/правки — не спамим справкой
    render = handle_message(msg["text"], cfg, archive)
    if render:
        tg_send(render["text"], chat_id=msg["chat"]["id"],
                reply_markup=render.get("markup"))


def _save_offset(offset):
    BOT_STATE_PATH.write_text(json.dumps({"offset": offset}, ensure_ascii=False),
                              encoding="utf-8")


def bot_poll(cfg):
    """Один ограниченный цикл опроса (--poll, резерв для GitHub Actions)."""
    if not TOKEN:
        sys.exit("❌ Нет TELEGRAM_BOT_TOKEN — бот не может опрашивать Telegram.")
    archive = load_json(ARCHIVE_PATH, {})
    offset = load_json(BOT_STATE_PATH, {}).get("offset", 0)
    window = cfg.get("bot_poll_seconds", 40)
    deadline = time.time() + window
    handled = 0
    while time.time() < deadline:
        try:
            updates = tg_get_updates(offset, timeout=min(20, window))
        except Exception as e:
            print(f"⚠️  getUpdates упал: {e}")
            break
        for up in updates:
            offset = up["update_id"] + 1
            try:
                process_update(up, cfg, archive)
            except Exception as e:
                print(f"⚠️  апдейт {offset}: {e}")
            handled += 1
        _save_offset(offset)
    print(f"🤖 Обработано сообщений: {handled}, offset={offset}")


def _git_pull_quiet():
    try:
        subprocess.run(["git", "-C", str(ROOT), "pull", "--quiet", "--no-rebase"],
                       timeout=60, capture_output=True)
    except Exception as e:
        print(f"⚠️  git pull: {e}")


def bot_serve(cfg):
    """Постоянный процесс long-polling: ответы за секунды (локально или на сервере)."""
    if not TOKEN:
        sys.exit("❌ Нет TELEGRAM_BOT_TOKEN — задай переменную окружения перед запуском.")
    print("🤖 Грант-радар: бот запущен (long-polling). Ctrl+C — остановить.")
    offset = load_json(BOT_STATE_PATH, {}).get("offset", 0)
    archive, mtime = {}, -1.0
    pull_every = int(cfg.get("serve_git_pull_seconds", 0))
    last_pull = 0.0
    while True:
        now = time.time()
        if pull_every and now - last_pull > pull_every:
            _git_pull_quiet()
            last_pull = now
        if ARCHIVE_PATH.exists() and ARCHIVE_PATH.stat().st_mtime != mtime:
            archive = load_json(ARCHIVE_PATH, {})
            mtime = ARCHIVE_PATH.stat().st_mtime
            print(f"📚 архив загружен: {len(archive)} записей")
        try:
            updates = tg_get_updates(offset, timeout=50)
        except Exception as e:
            print(f"⚠️  getUpdates: {e} — пауза 5 c")
            time.sleep(5)
            continue
        for up in updates:
            offset = up["update_id"] + 1
            try:
                process_update(up, cfg, archive)
            except Exception as e:
                print(f"⚠️  апдейт {offset}: {e}")
        if updates:
            _save_offset(offset)


def main():
    args = sys.argv[1:]

    cfg = load_json(CONFIG_PATH, None)
    if cfg is None:
        sys.exit("❌ Нет config.json")

    if "--flush" in args:
        flush_digest(cfg)
        return
    if "--serve" in args:
        bot_serve(cfg)
        return
    if "--poll" in args or "--bot" in args:
        bot_poll(cfg)
        return
    if "--report" in args:
        cmd_report(cfg)
        return
    if "--import" in args:
        i = args.index("--import")
        paths = args[i + 1:]
        if not paths:
            sys.exit("Использование: python radar.py --import <файл.csv|.json|.xlsx | папка>")
        cmd_import(cfg, paths)
        return
    if "--search" in args:
        i = args.index("--search")
        query = " ".join(args[i + 1:]).strip()
        if not query:
            sys.exit("Использование: python radar.py --search <тема>")
        cmd_search(cfg, query)
        return

    mode = str(cfg.get("mode", "stream")).lower()

    seen = load_json(SEEN_PATH, None)
    first_run = seen is None
    seen = seen or {}

    archive = load_json(ARCHIVE_PATH, {})
    today = today_iso()

    found = scan(cfg)
    new_items = [(s, it) for (s, it) in found if it["key"] not in seen]

    # Запоминаем всё найденное сразу (даже если не отправим/отложим)
    # и одновременно копим исторический архив для поиска.
    for src, it in found:
        seen[it["key"]] = it["title"]
        archive_upsert(archive, src, it, cfg, today)

    if first_run:
        extra = (
            " Коплю их и пришлю одним дайджестом по расписанию."
            if mode == "digest"
            else " Новые уведомления пойдут со следующих проверок."
        )
        tg_send(
            "🛰 <b>Грант-радар запущен!</b>\n"
            f"Слежу за источниками ({len(cfg['sources'])} шт.) в режиме "
            f"<b>{esc(mode)}</b>.\n"
            f"Сейчас в поле зрения — {len(found)} актуальных позиций."
            + extra
        )
    elif mode == "digest":
        # Копим новые находки в очередь — отправит отдельный запуск --flush.
        pending = load_json(PENDING_PATH, [])
        have = {p["key"] for p in pending}
        added = 0
        for src, it in new_items:
            if it["key"] in have:
                continue
            have.add(it["key"])
            pending.append(
                {
                    "key": it["key"],
                    "title": it["title"],
                    "link": it["link"],
                    "name": src["name"],
                    "emoji": src.get("emoji", "📌"),
                }
            )
            added += 1
        PENDING_PATH.write_text(
            json.dumps(pending, ensure_ascii=False, indent=1), encoding="utf-8"
        )
        print(f"🗂 Отложено в дайджест: +{added}, всего в очереди {len(pending)}")
    else:
        cap = cfg.get("max_alerts_per_run", 25)
        to_send = new_items[:cap]
        for src, it in to_send:
            tg_send(stream_item_msg(src, it))
            time.sleep(0.4)
        if len(new_items) > cap:
            tg_send(f"…и ещё {len(new_items) - cap} новых — покажу в следующий раз.")
        print(f"✅ Новых: {len(new_items)}, отправлено: {len(to_send)}")

    SEEN_PATH.write_text(
        json.dumps(seen, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    ARCHIVE_PATH.write_text(
        json.dumps(archive, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    print(f"💾 seen.json: {len(seen)} записей · archive.json: {len(archive)}")


if __name__ == "__main__":
    main()
